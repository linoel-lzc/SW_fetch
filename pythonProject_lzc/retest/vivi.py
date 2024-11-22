import re
import time
import openpyxl as op
from openpyxl.styles import Alignment

from netmiko import ConnectHandler



def telnet_to_switch_with_netmiko(host, username, password, enable_password):
    try:
        # 定义设备参数
        device = {
            "device_type": "cisco_ios_telnet",  # 使用 Telnet 连接 Cisco IOS
            "host": host,
            "username": username,
            "password": password,
            "secret": enable_password,  # enable 密码
        }

        # 建立连接
        connection = ConnectHandler(**device)

        # 进入特权模式
        connection.enable()


        return connection

    except Exception as e:
        return f"An error occurred: {e}"
def test1(c, start, end):
    lab = []
    for i in range(start, end+1):
        coo = f'show cdp neighbors gi 1/0/{i} detail'
        result = c.send_command(coo)
        # print(result)
        if 'Total cdp entries displayed : 0' in result:
            ip_address = ['0']
            mac = ['0']
            Platform = ['0']
            dict_list = {"id": f'gi 1/0/{i}', "ip_address": ip_address[0], "mac_address": mac[0],
                            "Platform": Platform[0]}
            lab.append(dict_list)
        else:
            ip_address = re.findall(r'\d+\.\d+\.\d+\.\d+', result)
            if len(ip_address) == 0:
                ip_address = ['0']

            mac = re.findall(r'SE[A-Z0-9]+', result)
            if len(mac) == 0:
                mac = ['0']

            Platform = re.findall(r'Platform: (.+?),', result)
            if len(Platform) == 0:
                Platform = ['0']


            dict_list = {"id": f'gi 1/0/{i}', "ip_address": ip_address[0], "mac_address": mac[0][3:15],
                            "Platform": Platform[0]}
            lab.append(dict_list)
    return lab
def Op_ToExcel(data, fileName):  # openpyxl库储存数据到excel
    list_column = ['A', 'B', 'C', 'D']
    wb = op.Workbook()  # 创建工作簿对象
    ws = wb['Sheet']  # 创建子表
    ws.append(['id', 'ip_address', 'mac_address', 'Platform'])  # 添加表头
    for i in range(len(list_column)):  # 设置列宽
        ws.column_dimensions[f'{list_column[i]}'].width = 20

    for i in range(len(data)):  # 添加数据
        d = data[i]["id"], data[i]["ip_address"], data[i]["mac_address"], data[i]["Platform"],
        ws.append(d)  # 每次写入一行

    max_rows = ws.max_row  # 获取最大行
    max_columns = ws.max_column  # 获取最大列
    align = Alignment(horizontal='center', vertical='center')
    # openpyxl的下标从1开始
    for i in range(1, max_rows + 1):  # 数据居中处理
        for j in range(1, max_columns + 1):
            ws.cell(i, j).alignment = align

    wb.save(fileName)

    wb.close()


if __name__ == "__main__":
    start = time.time()
    host = "10.74.62.99"
    username = "cisco"
    password = "cisco"
    enable_password = "cisco"
    connection = telnet_to_switch_with_netmiko(host, username, password, enable_password)

    list1 = test1(connection,1,8)
    print(list1)
    # Op_ToExcel(list1, 'a.xlsx')

    # 断开连接
    connection.disconnect()
    end = time.time()
    print(f'运行时间为{end-start}')

